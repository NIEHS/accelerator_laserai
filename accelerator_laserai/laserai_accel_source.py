"""Accelerator source component for LaserAI spreadsheet exports."""

from __future__ import annotations

from collections import OrderedDict
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import load_workbook

from accelerator_core.utils.xcom_utils import XcomPropsResolver
from accelerator_core.workflow.accel_source_ingest import (
    AccelIngestComponent,
    IngestPayload,
    IngestSourceDescriptor,
)


def _clean(value: Any) -> Any:
    """Trim spreadsheet text while retaining non-text scalar values."""
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _unique(values: Iterable[Any]) -> list[Any]:
    """Return non-empty values once, preserving their source order."""
    result = []
    for value in values:
        value = _clean(value)
        if value is not None and value not in result:
            result.append(value)
    return result


def _column_values(rows: list[Mapping[str, Any]], *headers: str) -> list[Any]:
    values = []
    for row in rows:
        for header in headers:
            if header in row:
                values.append(row[header])
    return _unique(values)


def _first_value(rows: list[Mapping[str, Any]], header: str) -> Any:
    values = _column_values(rows, header)
    return values[0] if values else None


def _rollups(
    rows: list[Mapping[str, Any]],
    headers_by_level: Mapping[str, str],
) -> list[dict[str, Any]]:
    """Keep each spreadsheet row's hierarchical values in one JSON entry."""
    rollups = []
    for row in rows:
        rollup = {
            f"level{level}": _clean(row.get(header)) or ""
            for level, header in headers_by_level.items()
        }
        if any(rollup.values()):
            rollups.append(rollup)
    return rollups


def _levels(
    rows: list[Mapping[str, Any]],
    headers_by_level: Mapping[str, str],
) -> dict[str, list[Any]]:
    """Group non-rollup hierarchical fields under JSON level keys."""
    return {
        str(level): _column_values(rows, header)
        for level, header in headers_by_level.items()
    }


def _raw_values(rows: list[Mapping[str, Any]]) -> dict[str, list[Any]]:
    result: OrderedDict[str, list[Any]] = OrderedDict()
    for row in rows:
        for header, value in row.items():
            value = _clean(value)
            if value is not None:
                result.setdefault(header, []).append(value)
    return dict(result)


def _record(reference_number: str, rows: list[Mapping[str, Any]]) -> dict[str, Any]:
    """Build the source-neutral record consumed by the later HEW crosswalk."""
    return {
        "source": "laserai",
        "source_reference_number": reference_number,
        "bibliographic": {
            "first_author": _first_value(rows, "1st Author"),
            "year": _first_value(rows, "Year"),
            "title": _first_value(rows, "Title"),
            "doi": _first_value(rows, "DOI"),
            "accession_number": _first_value(rows, "Accession Number"),
            "study_identifiers": _column_values(rows, "Study identifier"),
        },
        "review": {
            "reference_type": _first_value(rows, "Reference Type (extracted)"),
            "reference_type_comments": _column_values(
                rows, "Reference Type (comment)"
            ),
            "information_source": _first_value(rows, "Information Source (extracted)"),
            "information_source_comments": _column_values(
                rows, "Information Source (comment)"
            ),
            "recommend_for_removal": _first_value(
                rows, "Recommend for Removal (extracted)"
            ),
            "recommend_for_removal_comments": _column_values(
                rows, "Recommend for Removal (comment)"
            ),
            "postpone": _first_value(rows, "Postpone (extracted)"),
            "postpone_comments": _column_values(rows, "Postpone (comment)"),
        },
        "study_objectives": _column_values(
            rows, "Study objective (AI-generated) (extracted)"
        ),
        "exposures": _rollups(
            rows,
            {
                "1": "Exposure L1 (extracted)",
                "2": "Exposure L2 (extracted)",
                "3": "Exposure L3 (extracted)",
            },
        )
        + [
            {"level1": value, "level2": "", "level3": ""}
            for value in _column_values(rows, "Write-in (extracted)")
        ],
        "health_impacts": _rollups(
            rows,
            {
                "1": "Health Impact L1 (extracted)",
                "2": "Health Impact L2 (extracted)",
                "3": "Health Impact L3 (extracted)",
            },
        )
        + [
            {"level1": value, "level2": "", "level3": ""}
            for value in _column_values(rows, "Write-In (extracted)6")
        ],
        "geography": _rollups(
            rows,
            {
                "1": "Location L1 (extracted)",
                "2": "Location L2 (extracted)",
                "3": "Location L3 (extracted)",
            },
        )
        + [
            {"level1": value, "level2": "", "level3": ""}
            for value in _column_values(rows, "Write-in (extracted)8")
        ],
        "geographic_features": _column_values(
            rows, "Geographic Feature (extracted)"
        ),
        "data_and_models": {
            "data_resource_types": {
                "levels": _levels(
                    rows,
                    {
                        "1": "Data Resource Type L1 (extracted)",
                        "2": "Data Resource Type L2 (extracted)",
                    },
                ),
                "write_in": _column_values(rows, "Write-in (extracted)10"),
            },
            "model_types": _column_values(rows, "Model Type (extracted)"),
        },
        "special_topics": {
            "levels": _levels(
                rows,
                {
                    "1": "Special Topics L1 (extracted)",
                    "2": "Special Topics L2 (extracted)",
                },
            ),
            "write_in": _column_values(rows, "Write-in (extracted)12"),
        },
        # Keep the original sparse values, including repeated values and sentinels
        # such as "not reported", for audit and crosswalk decisions.
        "raw_values": _raw_values(rows),
    }


class LaserAIAccelSource(AccelIngestComponent):
    """Read a LaserAI export and emit one intermediate record per reference."""

    def __init__(
        self,
        ingest_source_descriptor: IngestSourceDescriptor,
        xcom_props_resolver: XcomPropsResolver,
    ):
        super().__init__(ingest_source_descriptor, xcom_props_resolver)

    @staticmethod
    def parse_workbook(identifier: str | Path) -> list[dict[str, Any]]:
        """Parse and aggregate a LaserAI workbook without creating payload objects."""
        workbook_path = Path(identifier)
        if not workbook_path.is_file():
            raise FileNotFoundError(f"LaserAI workbook not found: {workbook_path}")

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration as exc:
                raise ValueError("LaserAI workbook is empty") from exc

            headers = [_clean(header) for header in header_row]
            if "Reference number" not in headers:
                raise ValueError("LaserAI workbook must contain 'Reference number'")

            grouped: OrderedDict[str, list[dict[str, Any]]] = OrderedDict()
            reference_index = headers.index("Reference number")
            for row_number, values in enumerate(rows, start=2):
                row = {
                    header: value
                    for header, value in zip(headers, values)
                    if header is not None
                }
                if not any(_clean(value) is not None for value in row.values()):
                    continue

                reference_number = _clean(values[reference_index])
                if reference_number is None:
                    raise ValueError(
                        f"LaserAI row {row_number} contains data but no Reference number"
                    )
                reference_key = str(reference_number)
                grouped.setdefault(reference_key, []).append(row)

            return [
                _record(reference_number, reference_rows)
                for reference_number, reference_rows in grouped.items()
            ]
        finally:
            workbook.close()

    def ingest_single(
        self, identifier: str, additional_parameters: dict
    ) -> IngestPayload:
        """Ingest a workbook path supplied as the source identifier."""
        records = self.parse_workbook(identifier)
        ingest_payload = IngestPayload(self.ingest_source_descriptor)
        for record in records:
            self.report_individual(
                ingest_payload,
                record["source_reference_number"],
                record,
            )
        ingest_payload.ingest_successful = True
        return ingest_payload
